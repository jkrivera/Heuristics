# -*- coding: utf-8 -*-


from openpyxl import Workbook
from openpyxl import load_workbook


import numpy as np
import math

def reading(ins):

    wb = load_workbook(filename="Datos.xlsx")
    
    sheet = wb["I" + str(ins)]
    
    N = sheet["A1"].value
    M = sheet["B1"].value
    P = sheet["C1"].value
    
    c = np.zeros((P,N))
    for p in range(P):
        for j in range(N):
            c[p][j] = sheet.cell(row=2+M+p, column=1+j).value
    
    a = np.zeros((M,N))
    for i in range(M):
        for j in range(N):
            a[i][j] = sheet.cell(row=2+i, column=1+j).value

    b = np.zeros((M))
    for i in range(M):
        b[i] = sheet.cell(row=2+i, column=1+N).value
        
    P = 1

    
    return N, M, P, c, a, b