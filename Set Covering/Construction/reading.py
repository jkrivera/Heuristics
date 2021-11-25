# -*- coding: utf-8 -*-


from openpyxl import Workbook
from openpyxl import load_workbook


import numpy as np
import math

def reading(ins):

    wb = load_workbook(filename="Data.xlsx")
    
    sheet = wb[ins]
    
    m = sheet["A1"].value
    n = sheet["B1"].value
    K = sheet["C1"].value
    
    c = np.zeros((n))
    for i in range(n):
        c[i] = sheet.cell(row=2, column=1+i).value
    
    a = np.zeros((n,m))
    ni = np.zeros((n))
    nj = np.zeros((m))
    listi = np.zeros((n,m))
    listj = np.zeros((m,n))
    w = np.zeros((m))
    for j in range(m):
        nj[j] = sheet.cell(row=3+j, column=1).value
        for i in range(int(nj[j])):
            k = sheet.cell(row=3+j, column=2+i).value
            a[k-1][j] = 1
            listj[j][i] = int(k-1)
            ni[k-1] = ni[k-1]+1
            listi[k-1][int(ni[k-1])-1] = j
        w[j] = sheet.cell(row=3+j, column=2+int(nj[j])).value

    return n, m, c, a, ni, nj, listi, listj, K, w